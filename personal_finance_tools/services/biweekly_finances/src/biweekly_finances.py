import calendar
import json
import os
import shutil
import sys
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

from openpyxl import load_workbook


class BiweeklyFinanceProcessor:
    """
    Class for the management and creation of biweekly finance Excel Files
    """

    def __init__(self, year, month):
        self.month_dt = datetime(year=year, month=month, day=1)
        self.previous_month_dt = self.month_dt - relativedelta(months=1)
        self.service_config = "../service_config.json"
        self.template_file = self._get_config_value(['templateFilePath'])
        self.output_base_dir = self._get_config_value(['outputBaseDir'])

    def create_new_excel_file(self):
        print("template_file: {0}".format(self.template_file))
        print("output_base_dir: {0}".format(self.output_base_dir))

        # Create Directories if they don't exist
        create_dir_if_not_exists(self.output_base_dir)  # Base Directory
        year_dir = os.path.join(self.output_base_dir, str(self.month_dt.year))  # Year Directory
        create_dir_if_not_exists(year_dir)

        # Copy Template File and Rename
        new_excel_filename = "{0} - {1}.xlsx".format(str(self.month_dt.year),
                                                     calendar.month_name[self.month_dt.month])
        output_excel_filepath = self._copy_and_rename_template(dest_dir=year_dir, new_file_name=new_excel_filename)

        # Determine Previous Month's filepath
        previous_excel_filename = "{0} - {1}.xlsx".format(str(self.previous_month_dt.year),
                                                          self.previous_month_dt.month)
        previous_month_file_path = os.path.join(output_excel_filepath, str(self.month_dt.year), previous_excel_filename)

        # If It Exists, read previous month's Excel File and copy values
        self._populate_excel_values_from_previous_month(new_file_path=output_excel_filepath,
                                                        previous_file_path=previous_month_file_path)

    def _populate_excel_values_from_previous_month(self, new_file_path: str, previous_file_path: str):
        """
            Populates all values of the biweekly finance file as a successor to the previous file
            :param new_file_path: new biweekly finance file to populate with values
            :param previous_file_path: file path of biweekly finance file of the previous month
            """
        # New File
        new_workbook = load_workbook(new_file_path)
        new_workbook_datetime = determine_biweekly_finance_file_datetime(new_file_path)
        new_month_abbr = new_workbook_datetime.strftime('%b')

        # Rename New File Worksheet Names
        print("Renaming worksheets...")
        new_workbook["Pay 1"].title = "{0} - Pay 1".format(new_month_abbr)
        new_workbook["Pay 2"].title = "{0} - Pay 2".format(new_month_abbr)
        new_workbook["Pay 3"].title = "{0} - Pay 3".format(new_month_abbr)
        new_workbook.save(new_file_path)

        # Set Variables for new Workbook and Worksheets
        new_workbook = load_workbook(new_file_path)
        new_workbook_pay_1_worksheet = new_workbook["{0} - Pay 1".format(new_month_abbr)]
        new_workbook_pay_2_worksheet = new_workbook["{0} - Pay 2".format(new_month_abbr)]
        new_workbook_pay_3_worksheet = new_workbook["{0} - Pay 3".format(new_month_abbr)]
        new_workbook_receipt_1_worksheet = new_workbook["Receipt 1"]
        new_workbook_receipt_2_worksheet = new_workbook["Receipt 2"]
        new_workbook_receipt_3_worksheet = new_workbook["Receipt 3"]

        # Set static variables for biweekly finance excel cell, column, and row values
        pay_start_date_cell = "B25"
        pay_end_date_cell = "B26"
        receipt_start_date_cell = "B24"
        receipt_end_date_cell = "B25"
        total_amount_column_letter = "C"
        new_total_amount_column_letter = "H"
        money_category_start_row_inclusive = 2
        money_category_end_row_exclusive = 22

        # Pulling Values
        if os.path.isfile(previous_file_path):  # Previous Month's workbook is found
            previous_workbook = load_workbook(previous_file_path, data_only=True)
            previous_workbook_datetime = determine_biweekly_finance_file_datetime(previous_file_path)
            previous_month_abbr = previous_workbook_datetime.strftime('%b')

            # Previous Workbook most recent worksheet
            print("Determining most recent pay worksheet...")
            if "{0} - Pay 3".format(previous_month_abbr) in previous_workbook.sheetnames:
                # Use Pay 3
                previous_workbook_most_recent_pay_worksheet = previous_workbook[
                    "{0} - Pay 3".format(previous_month_abbr)]
            elif "{0} - Pay 2".format(previous_month_abbr) in previous_workbook.sheetnames:
                # Use Pay 2
                previous_workbook_most_recent_pay_worksheet = previous_workbook[
                    "{0} - Pay 2".format(previous_month_abbr)]
            else:
                # Use Pay 1
                previous_workbook_most_recent_pay_worksheet = previous_workbook[
                    "{0} - Pay 1".format(previous_month_abbr)]
            previous_workbook_last_date = previous_workbook_most_recent_pay_worksheet[pay_end_date_cell].value

            # Paste old workbook "New Total Amount" in new workbook's "Total Amount" for each category
            print("Calculating category money values...")
            for row in range(money_category_start_row_inclusive, money_category_end_row_exclusive):
                src_cell = new_total_amount_column_letter + str(row)
                dst_cell = total_amount_column_letter + str(row)
                value = previous_workbook_most_recent_pay_worksheet[src_cell].value
                value2 = previous_workbook_most_recent_pay_worksheet[src_cell]
                new_workbook_pay_1_worksheet[dst_cell].value = value
            new_workbook.save(new_file_path)

        else:  # No previous file
            first_day = int(input("No workbook found for previous month to determine start date of the new workbook. "
                                  "\nInput start day of this workbook: "))
            first_month = new_workbook_datetime.month
            first_year = new_workbook_datetime.year
            new_workbook_first_date = datetime(month=first_month, year=first_year, day=first_day)
            previous_workbook_last_date = new_workbook_first_date - timedelta(days=1)

        # Calculate Dates
        print("Calculating dates for new worksheets...")
        new_biweekly_1_start_date = previous_workbook_last_date + timedelta(days=1)
        new_biweekly_1_end_date = new_biweekly_1_start_date + timedelta(days=13)
        new_biweekly_2_start_date = new_biweekly_1_end_date + timedelta(days=1)
        new_biweekly_2_end_date = new_biweekly_2_start_date + timedelta(days=13)
        new_biweekly_3_start_date = new_biweekly_2_end_date + timedelta(days=1)
        new_biweekly_3_end_date = new_biweekly_3_start_date + timedelta(days=13)

        # Paste new workbook dates
        new_workbook_pay_1_worksheet[pay_start_date_cell] = new_biweekly_1_start_date
        new_workbook_pay_1_worksheet[pay_end_date_cell] = new_biweekly_1_end_date
        new_workbook_receipt_1_worksheet[receipt_start_date_cell] = new_biweekly_1_start_date
        new_workbook_receipt_1_worksheet[receipt_end_date_cell] = new_biweekly_1_end_date
        new_workbook_pay_2_worksheet[pay_start_date_cell] = new_biweekly_2_start_date
        new_workbook_pay_2_worksheet[pay_end_date_cell] = new_biweekly_2_end_date
        new_workbook_receipt_2_worksheet[receipt_start_date_cell] = new_biweekly_2_start_date
        new_workbook_receipt_2_worksheet[receipt_end_date_cell] = new_biweekly_2_end_date
        new_workbook_pay_3_worksheet[pay_start_date_cell] = new_biweekly_3_start_date
        new_workbook_pay_3_worksheet[pay_end_date_cell] = new_biweekly_3_end_date
        new_workbook_receipt_3_worksheet[receipt_start_date_cell] = new_biweekly_3_start_date
        new_workbook_receipt_3_worksheet[receipt_end_date_cell] = new_biweekly_3_end_date
        new_workbook.save(new_file_path)

        # Update "Total Amount" formulas to point to previous pay in same workbook instead of template workbook
        for row in range(money_category_start_row_inclusive, money_category_end_row_exclusive):
            cell = total_amount_column_letter + str(row)
            pay_2_cell_value = new_workbook_pay_2_worksheet[cell].value.replace('Pay 1', f'{new_month_abbr} - Pay 1')
            pay_3_cell_value = new_workbook_pay_3_worksheet[cell].value.replace('Pay 2', f'{new_month_abbr} - Pay 2')
            new_workbook_pay_2_worksheet[cell].value = pay_2_cell_value
            new_workbook_pay_3_worksheet[cell].value = pay_3_cell_value
        new_workbook.save(new_file_path)

        # Remove Pay 3 and Receipt 3 if it should not be in the new workbook (if end date is into next month)
        _, num_of_days_in_month = calendar.monthrange(new_workbook_datetime.year, new_workbook_datetime.month)
        last_day_of_month = datetime(new_workbook_datetime.year, new_workbook_datetime.month, num_of_days_in_month)
        if new_biweekly_3_end_date > last_day_of_month:
            print("3rd pay end date is into next month. Removing 3rd Pay from this worksheet...")
            new_workbook.remove(new_workbook_pay_3_worksheet)
            new_workbook.remove(new_workbook_receipt_3_worksheet)
            new_workbook.save(new_file_path)
        if new_biweekly_2_end_date > last_day_of_month:
            print("2nd pay end date is into next month. Removing 2nd Pay from this worksheet...")
            new_workbook.remove(new_workbook_pay_2_worksheet)
            new_workbook.remove(new_workbook_receipt_2_worksheet)
            new_workbook.save(new_file_path)

    def _copy_and_rename_template(self, dest_dir: str, new_file_name: str):
        try:
            # Copy file
            new_file_copy = shutil.copy(self.template_file, dest_dir)
            # Rename File
            new_file_path = os.path.join(dest_dir, new_file_name)
            os.rename(src=new_file_copy, dst=new_file_path)
            print("Copied and renamed template file: {0}".format(new_file_path))
            return new_file_path
        except Exception as e:
            print("Failed to copy or rename file.")
            print(e)

    def _get_config_value(self, key_list=None):
        """
        Returns the value from the service_config json file
        :param key_list: list of keys, each key being a subkey of the key before it
        :return: value from config file
        """
        # Default Value (Pulls all config)
        if key_list is None:
            key_list = []

        # Open File
        service_config_file = open(self.service_config)
        config_json = json.load(service_config_file)

        # Search for Key
        current_value = config_json
        try:
            for key in key_list:
                current_value = current_value[key]
        except KeyError as e:
            print("Unable to find specified config key: {0}".format(key_list))
            print(e)

        # Close File and return value
        service_config_file.close()
        return current_value


#####################################
# STATIC METHODS
#####################################
def create_dir_if_not_exists(dir_path: str):
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)


def determine_biweekly_finance_file_datetime(file_path: str):
    """
    Returns a datetime object of the biweekly finance file month and year
    :param file_path: biweekly finance .xlsx file
    :return: datetime month and year of the file
    :rtype: datetime
    """
    file_name = os.path.basename(file_path)
    return datetime.strptime(file_name, '%Y - %B.xlsx')


#####################################
# MAIN METHOD
#####################################
def main():
    # Expected Arguments
    expected_argument_length = 3
    received_argument_length = len(sys.argv)
    if received_argument_length != expected_argument_length:
        raise ValueError("Expected Arguments: {0} | Received Arguments: {1}".format(expected_argument_length,
                                                                                    received_argument_length))

    # Get Input Arguments
    new_file_year = int(sys.argv[1])
    new_file_month = int(sys.argv[2])

    # Create new file based on template
    biweekly_finance_processor = BiweeklyFinanceProcessor(year=new_file_year, month=new_file_month)
    biweekly_finance_processor.create_new_excel_file()


if __name__ == "__main__":
    main()
